from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import random
import os


# ==========================
# 参数
# ==========================

subjects = {
    "语文": 150,
    "数学": 150,
    "英语": 150,
    "物理": 100,
    "化学": 100,
    "生物": 100,
    "历史": 100,
    "政治": 100,
    "地理": 100
}


student_count = 50


names = [
    "张伟", "李娜", "王强", "赵敏",
    "刘洋", "陈晨", "杨帆",
    "黄磊", "周杰", "吴迪",
    "徐静", "孙浩"
]


save_dir = "成绩数据"

os.makedirs(
    save_dir,
    exist_ok=True
)



# ==========================
# 样式
# ==========================

def format_sheet(ws):

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="4F81BD"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    for col in ws.columns:

        length = max(
            len(str(c.value))
            if c.value else 0
            for c in col
        )

        ws.column_dimensions[
            col[0].column_letter
        ].width = length + 3



# ==========================
# 成绩生成
# ==========================

def random_total(full):

    """
    模拟考试成绩
    """

    score = random.gauss(
        full * 0.75,
        full * 0.15
    )

    return int(
        max(
            full*0.3,
            min(score, full)
        )
    )



def split_questions(total):

    """
    总分拆成23题
    """

    weights = [
        random.random()
        for _ in range(23)
    ]

    s = sum(weights)

    result = []

    remain = total


    for i,w in enumerate(weights):

        if i == 22:
            score = remain

        else:
            score = round(
                total*w/s
            )

            remain -= score

        result.append(score)


    return result



# ==========================
# 学生名单
# ==========================

students = []

for i in range(student_count):

    students.append(
        {
            "id":2026001+i,
            "name":
                random.choice(names)
                + str(i+1)
        }
    )



# 保存总成绩

summary = {}

for stu in students:

    summary[stu["id"]] = {
        "学生":stu["name"]
    }



# ==========================
# 生成各学科文件
# ==========================

for subject, full_score in subjects.items():


    wb = Workbook()

    ws = wb.active

    ws.title = subject



    ws.append(
        [
            "学号",
            "学生",
            *[
                f"小题{i}"
                for i in range(1,24)
            ],
            "总分"
        ]
    )


    for stu in students:


        total = random_total(
            full_score
        )


        questions = split_questions(
            total
        )


        ws.append(
            [
                stu["id"],
                stu["name"],
                *questions,
                total
            ]
        )


        summary[
            stu["id"]
        ][subject] = total



    format_sheet(ws)


    wb.save(
        f"{save_dir}/{subject}成绩.xlsx"
    )



# ==========================
# 生成总成绩表
# ==========================


wb = Workbook()

ws = wb.active

ws.title = "总成绩"



headers = [
    "学号",
    "学生",
    *subjects.keys(),
    "总分"
]


ws.append(headers)



for sid,data in summary.items():

    total = sum(
        data[sub]
        for sub in subjects
    )


    ws.append(
        [
            sid,
            data["学生"],
            *[
                data[sub]
                for sub in subjects
            ],
            total
        ]
    )


format_sheet(ws)



wb.save(
    f"{save_dir}/总成绩汇总.xlsx"
)


print("全部成绩文件生成完成")